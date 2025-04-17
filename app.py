from gtts import gTTS
import os
import requests
import tempfile
import streamlit as st
import speech_recognition as sr
import whisper
from openpyxl import load_workbook
from datetime import datetime
import re

# ---------- CONFIG ----------
model = whisper.load_model("base")
TEMPLATE_PATH = "Stephen_Gould_Daily_Report_Template.xlsx"
REPORT_PATH = "VoiceReport_Streamlit.xlsx"

# ---------- SUPPORT ----------
NUM_WORDS = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4,
    "five": 5, "six": 6, "seven": 7, "eight": 8, "nine": 9,
    "ten": 10
}

def extract_number(text):
    cleaned = re.sub(r"[^\w\s]", "", text.lower())
    digits = re.findall(r'\d+', cleaned)
    if digits:
        return int(digits[0])
    for word in cleaned.split():
        if word in NUM_WORDS:
            return NUM_WORDS[word]
    return None

def get_ip_location():
    try:
        res = requests.get("http://ip-api.com/json/").json()
        city = res.get("city", "Unknown")
        region = res.get("regionName", "")
        country = res.get("country", "")
        lat = res.get("lat", None)
        lon = res.get("lon", None)
        return f"{city}, {region}, {country}", lat, lon
    except:
        return "Unknown location", None, None

def get_weather(lat, lon):
    api_key = "c9f90ac0f54001b124c4c0ab36ba0e70"
    try:
        res = requests.get(
            f"https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&appid={api_key}&units=imperial"
        ).json()
        desc = res["weather"][0]["description"].capitalize()
        temp = res["main"]["temp"]
        return f"{desc}, {temp}°F"
    except:
        return "Weather unavailable"

def speak(text):
    tts = gTTS(text=text, lang='en')
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".mp3")
    tts.save(temp_file.name)

    audio_path = temp_file.name
    audio_bytes = open(audio_path, "rb").read()

    # Display invisible audio player with autoplay via HTML
    st.markdown(
        f"""
        <audio autoplay>
            <source src="data:audio/mp3;base64,{audio_bytes.hex()}" type="audio/mp3">
        </audio>
        """,
        unsafe_allow_html=True
    )
    
    os.remove(temp_file.name)


def record_and_transcribe():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        st.info("🎧 Listening...")
        audio = r.listen(source, phrase_time_limit=8)
        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as temp_audio:
            temp_audio.write(audio.get_wav_data())
            result = model.transcribe(temp_audio.name)
        return result["text"].strip()

# ---------- UI START ----------
st.title("🎤 Voice-Driven Daily Report")

BASE_QUESTIONS = [
    ("What is the contract number?", "C4"),
    ("Who is the superintendent?", "C5"),
    ("What is the job name?", "F4"),
    ("What is the report number?", "K5"),
    ("How many hours were worked?", "B26"),
    ("Is there a weather impact? (Yes or No)", "J10"),
    ("How many Trident Builders labor entries?", None),
    ("How many Subcontractor labor entries?", None),
    ("How many visitor entries?", None),
    ("How many 'Work Performed This Date' entries?", None)
]

# ---------- Session State ----------
if "questions" not in st.session_state:
    st.session_state.questions = BASE_QUESTIONS.copy()
    st.session_state.answers = {}
    st.session_state.q_index = 0
    st.session_state.retry = False
    location, lat, lon = get_ip_location()
    weather = get_weather(lat, lon) if lat and lon else "Unavailable"
    st.session_state.location_str = location
    st.session_state.weather_str = weather

# ---------- Main Q&A Logic ----------
if st.session_state.q_index < len(st.session_state.questions):
    question, cell = st.session_state.questions[st.session_state.q_index]
    st.write(f"**Q{st.session_state.q_index + 1}:** {question}")

    speak(question)
    response = record_and_transcribe()
    st.success(f"You said: {response}")

    if cell:
        st.session_state.answers[cell] = response
        st.session_state.q_index += 1
        st.rerun()
    else:
        count = extract_number(response)
        if count is None:
            st.warning("⚠️ Couldn't understand the number. Repeating the question...")
            st.rerun()
        else:
            q = st.session_state.questions
            base = 15
            if "Trident Builders" in question:
                for i in range(count):
                    q.extend([(f"Position #{i+1}?", f"B{base+i}"),
                              (f"Crew size for Position #{i+1}?", f"C{base+i}"),
                              (f"Hours for Position #{i+1}?", f"D{base+i}")])
            elif "Subcontractor" in question:
                for i in range(count):
                    q.extend([(f"Company #{i+1}?", f"E{base+i}"),
                              (f"Crew size for Company #{i+1}?", f"F{base+i}"),
                              (f"Hours for Company #{i+1}?", f"G{base+i}")])
            elif "visitor" in question.lower():
                for i in range(count):
                    q.extend([(f"Visitor Company #{i+1}?", f"H{base+i}"),
                              (f"Visitor Name #{i+1}?", f"I{base+i}"),
                              (f"Visitor Hours #{i+1}?", f"J{base+i}")])
            elif "Work Performed" in question:
                for i in range(count):
                    idx = 32 + i
                    q.extend([(f"What is Work Performed #{i+1}?", f"B{idx}"),
                              (f"Equip/Tools on Site #{i+1}?", f"G{idx}"),
                              (f"Contractor for entry #{i+1}?", f"I{idx}")])
            st.session_state.q_index += 1
            st.rerun()

# ---------- Final Save ----------
else:
    st.success("✅ All questions complete!")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    for cell, val in st.session_state.answers.items():
        ws[cell] = val
    ws["F5"] = st.session_state.location_str
    ws["C8"] = st.session_state.weather_str
    ws["K4"] = datetime.today().strftime("%Y-%m-%d")
    wb.save(REPORT_PATH)
    with open(REPORT_PATH, "rb") as f:
        st.download_button("📥 Download Report", f, file_name="VoiceReport_Streamlit.xlsx")
